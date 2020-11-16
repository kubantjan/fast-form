import os
import tempfile
from typing import List, Tuple

import numpy as np
import openpyxl
import pandas as pd
from cv2 import cv2
from openpyxl import load_workbook

from fast_form.structure_parser.form_structure_dataclasses import FormStructure, FieldType

PATIENT_ID_FIELD = "patient_id"
SHEET_NAME = "questionnaire_data_output"


def output_data(form_data: FormStructure, maybe_patient_id: str) -> Tuple[pd.DataFrame, List[np.ndarray]]:
    form_dict = dict()
    form_images = []
    question_number = 0
    patient_id = maybe_patient_id
    for page_data in form_data.form_page_data.values():
        for field in page_data.fields:
            if field.name == PATIENT_ID_FIELD:
                assert isinstance(field.recognizing_results.recognized, str)
                read_patient_id = field.recognizing_results.recognized.replace(" ", "")
                if len(read_patient_id) > 0:
                    patient_id = read_patient_id

    for page_data in form_data.form_page_data.values():
        for field in page_data.fields:
            if field.name != PATIENT_ID_FIELD:
                field_dict = dict()
                field_dict["patient_id"] = patient_id
                field_dict['field_type'] = field.type.name
                field_dict['name'] = field.name
                field_dict['data'] = field.recognizing_results.recognized
                field_dict['acc'] = field.recognizing_results.accuracy
                form_images.append(field.img)
                form_dict[question_number] = field_dict
                question_number += 1

    df = pd.DataFrame.from_dict(form_dict, orient="index")
    return df, form_images


def save_data(df: pd.DataFrame, images: List[np.ndarray], excel_path: str):
    with tempfile.TemporaryDirectory() as temp_image_folder:

        if not os.path.exists(excel_path):
            df.to_excel(excel_path, index=False, sheet_name=SHEET_NAME)
            start_index = 1
            book = load_workbook(filename=excel_path)
            sheets_dict = {ws.title: ws for ws in book.worksheets}
        else:
            book = load_workbook(excel_path)
            writer = pd.ExcelWriter(excel_path, engine='openpyxl')
            writer.book = book
            sheets_dict = {ws.title: ws for ws in book.worksheets}
            writer.sheets = sheets_dict
            start_index = writer.sheets[SHEET_NAME].max_row
            df.to_excel(writer, sheet_name=SHEET_NAME, startrow=start_index, index=False, header=False)
            writer.save()
        sheet = sheets_dict[SHEET_NAME]
        for i, img, response, field_type in zip(df.index + start_index, images, df.data.values, df.field_type.values):
            if field_type != FieldType.SINGLE_CHOICE or response < 0:
                temp_image_path = os.path.join(temp_image_folder, f"{i}_img.png")
                height = 60
                width = int(height / img.shape[0] * img.shape[1])
                sheet.row_dimensions[i + 1].height = 50

                dim = (width, height)
                resized = cv2.resize(img, dim, interpolation=cv2.INTER_AREA)

                cv2.imwrite(temp_image_path, resized)
                img = openpyxl.drawing.image.Image(temp_image_path)
                img.anchor = f'F{i + 1}'
                sheet.add_image(img)
        book.save(excel_path)
