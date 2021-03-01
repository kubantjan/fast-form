import json
import logging
import os

import dacite
import openpyxl
import pandas as pd
from openpyxl import load_workbook

from fast_form.config.configuration_dataclasses import PathsForProcessingConfig, VALIDATION_EXCEL_NAME
from fast_form.config.configuration_loading import get_processing_config
from fast_form.outputting.process_document import process_document_and_add_to_validation_excel
from fast_form.structure_parser.form_structure_dataclasses import FieldType

SHEET_WITH_RESULTS = "automatic_results"

logger = logging.getLogger(__name__)


def process_to_validation_excel(paths_for_processing_config: PathsForProcessingConfig):
    processing_config = get_processing_config(paths_for_processing_config)

    if os.path.exists(paths_for_processing_config.validation_excel_path):
        os.remove(paths_for_processing_config.validation_excel_path)

    document_names = [file for file in os.listdir(paths_for_processing_config.folder_with_documents_path) if
                      file.endswith('.pdf') or file.endswith(".jpg") or file.endswith(".png")]
    logger.info(f"Processing to validation excel from {paths_for_processing_config.folder_with_documents_path}. Number"
                f" of documents is {len(document_names)}")
    for document_name in document_names:
        logging.debug(f"Processing document {document_name}")
        process_document_and_add_to_validation_excel(
            os.path.join(paths_for_processing_config.folder_with_documents_path, document_name),
            processing_config,
            paths_for_processing_config.validation_excel_path
        )
    logger.info(
        f"Successfully processed all documents from '{paths_for_processing_config.folder_with_documents_path}' to "
        f"validation excel '{paths_for_processing_config.validation_excel_path}'")


def process_to_final_excel(paths_for_processing_config: PathsForProcessingConfig):
    if not os.path.exists(paths_for_processing_config.validation_excel_path):
        raise FileNotFoundError(
            f"Validation excel with path '{paths_for_processing_config.validation_excel_path}' must exist")

    validation_df = pd.read_excel(paths_for_processing_config.validation_excel_path)
    one_patient_per_row_df = (
        validation_df
            .set_index(['patient_id', 'name'])
            .loc[lambda df: df.field_type == FieldType.SINGLE_CHOICE]
            .data
            .apply(lambda response_or_error: response_or_error if response_or_error >= 0 else "")
            .unstack(level=1)
    )

    combined_one_patient_per_row_df = one_patient_per_row_df

    if os.path.exists(paths_for_processing_config.final_excel_path):
        xl = pd.ExcelFile(paths_for_processing_config.final_excel_path)
        if SHEET_WITH_RESULTS in xl.sheet_names:  # see all sheet names
            old_one_patient_per_row_df = xl.parse(
                index_col="patient_id",
                keep_default_na=False,
                na_values=[],
                sheet_name=SHEET_WITH_RESULTS)

            combined_one_patient_per_row_df = one_patient_per_row_df.combine_first(old_one_patient_per_row_df)
        book = load_workbook(paths_for_processing_config.final_excel_path)
    else:
        book = openpyxl.Workbook()

    writer = pd.ExcelWriter(paths_for_processing_config.final_excel_path, engine='openpyxl')
    if SHEET_WITH_RESULTS in book.sheetnames:
        book.remove(book[SHEET_WITH_RESULTS])
    writer.book = book
    combined_one_patient_per_row_df.to_excel(writer, sheet_name=SHEET_WITH_RESULTS)
    writer.save()
    logger.info(f"Successfully processed from validation excel: '{paths_for_processing_config.validation_excel_path}'"
                f" to final excel '{paths_for_processing_config.final_excel_path}'")


def load_paths_for_processing_config(path_configuration: str) -> PathsForProcessingConfig:
    def join(rel_path):
        return os.path.join(os.path.dirname(path_configuration), rel_path)

    with open(path_configuration, "rb") as f:
        relative_paths = dacite.from_dict(data_class=PathsForProcessingConfig, data=json.loads(f.read()))
    return PathsForProcessingConfig(
        root=join("."),
        template_path=join(relative_paths.template_path),
        folder_with_documents_path=join(relative_paths.folder_with_documents_path),
        final_excel_path=join(relative_paths.final_excel_path),
        form_structure_path=join(relative_paths.form_structure_path),
        validation_excel_path=join(VALIDATION_EXCEL_NAME)
    )
