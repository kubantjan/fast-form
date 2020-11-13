import logging
import os

import pandas as pd
from openpyxl import load_workbook

from process_to_validation_excel import load_paths_for_processing_config, PATH_TO_PATH_CONFIG, VALIDATION_EXCEL_NAME

SHEET_WITH_RESULTS = "automatic_results"

logging.basicConfig(
    format='%(asctime)s %(levelname)-8s %(message)s',
    level=logging.INFO,
    datefmt='%Y-%m-%d %H:%M:%S')

if __name__ == '__main__':
    paths_for_processing_config = load_paths_for_processing_config(PATH_TO_PATH_CONFIG)
    validation_excel_path = os.path.join(paths_for_processing_config.folder_with_documents_path,
                                         VALIDATION_EXCEL_NAME)
    if not os.path.exists(validation_excel_path):
        raise FileNotFoundError(f"Validation excel with path {validation_excel_path} must exist")

    validation_df = pd.read_excel(validation_excel_path)
    one_patient_per_row_df = (
        validation_df
            .set_index(['patient_id', 'name'])
            .data
            .apply(lambda response_or_error: response_or_error if response_or_error >= 0 else "")
            .unstack(level=1)
    )

    combined_one_patient_per_row_df = one_patient_per_row_df

    if os.path.exists(paths_for_processing_config.final_excel_path):
        xl = pd.ExcelFile(paths_for_processing_config.final_excel_path)
        if SHEET_WITH_RESULTS in xl.sheet_names:  # see all sheet names
            old_one_patient_per_row_df = xl.parse(
                index_col="patient_id",
                keep_default_na=False,
                na_values=[],
                sheet_name=SHEET_WITH_RESULTS)

            combined_one_patient_per_row_df = one_patient_per_row_df.combine_first(old_one_patient_per_row_df)

    book = load_workbook(paths_for_processing_config.final_excel_path)
    writer = pd.ExcelWriter(paths_for_processing_config.final_excel_path, engine='openpyxl')
    if SHEET_WITH_RESULTS in book.sheetnames:
            book.remove(book[SHEET_WITH_RESULTS])
    writer.book = book
    combined_one_patient_per_row_df.to_excel(writer, sheet_name=SHEET_WITH_RESULTS)
    writer.save()
